import logging
from unittest import result
import pandas as pd
from django.http import HttpResponse
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.exceptions import bad_request
from django.forms.models import model_to_dict
from api.models import Category, Post
from django.db import transaction


def get_xlsx_report(posts, categories):
    post_df = pd.DataFrame(posts)[["category_id", "title", "description", "price", "weight", "country"]]
    cat_df = pd.DataFrame(categories)[["id", "name", "description"]]
    result = pd.merge(
        left=post_df,
        right=cat_df,
        left_on="category_id",
        right_on="id",
        how="inner",
        suffixes=("_post", "_category")
    )
    result = result[["name", "description_category", "title", "description_post", "price", "weight", "country"]]
    result.columns = ["category_name", "category_description", "post_title", "post_description", "price", "weight", "country"]
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        result.to_excel(
            writer,
            index=False,
            sheet_name="Текущие посты"
        )
        workbook = writer.book
        worksheet = writer.sheets["Текущие посты"]
        header_font = Font(bold=True)
        for col_num, column_title in enumerate(result.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[col_letter].width = length + 4
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    return response


def add_posts_file(data, request=None):
    df = pd.read_excel(data)
    posts = df.to_dict(orient="records")
    answer = []
    errors = []
    for i, post in enumerate(posts, start=1):
        try:
            category_name = post.get("category_name")
            category_desc = post.get("category_description", "")
            category = Category.objects.filter(name=category_name).first()
            if not category:
                category = Category.objects.create(
                    name=category_name,
                    description=category_desc
                )
            new_post = Post.objects.create(
                category=category,
                title=post.get("post_title"),
                description=post.get("post_description"),
                price=post.get("price", 0),
                weight=post.get("weight", 0),
                country=post.get("country", "")
            )
            answer.append(model_to_dict(new_post))
        except Exception as e:
            logging.error(f"Ошибка на строке {i}: {e}")
            errors.append({"row": i, "error": str(e)})
    return {"success": answer, "errors": errors}


def replace_posts_file(data, request):
    df = pd.read_excel(data)
    posts = df.to_dict(orient="records")
    answer = []
    errors = []

    try:
        with transaction.atomic():
            Post.objects.all().delete()
            Category.objects.all().delete()

            for i, post in enumerate(posts, start=1):
                try:
                    category_name = post.get("category_name")
                    category_desc = post.get("category_description", "")
                    category = Category.objects.filter(name=category_name).first()
                    if not category:
                        category = Category.objects.create(
                            name=category_name,
                            description=category_desc
                        )

                    new_post = Post.objects.create(
                        category=category,
                        title=post.get("post_title"),
                        description=post.get("post_description"),
                        price=post.get("price", 0),
                        weight=post.get("weight", 0),
                        country=post.get("country", "")
                    )
                    answer.append(model_to_dict(new_post))
                except Exception as e:
                    logging.error(f"Ошибка на строке {i}: {e}")
                    errors.append({"row": i, "error": str(e)})
            if errors:
                raise ValueError("Есть ошибки при импорте")
    except Exception as e:
        return {"success": [], "errors": errors or [{"error": str(e)}]}
    return {"success": answer, "errors": errors}